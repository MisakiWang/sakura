import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from copy import copy
import os

# ---------- 中文大写金额转换函数（支持负数，加“负”字） ----------
def digit_to_chinese(num):
    # 确保是数字
    if not isinstance(num, (int, float)):
        try:
            num = float(num)
        except:
            num = 0.0
    
    # 判断是否为负数（非零负数）
    is_negative = num < 0
    num = abs(num)          # 按绝对值处理
    num = round(num, 2)
    
    integer_part = int(num)
    decimal_part = int(round((num - integer_part) * 100))

    chinese_digits = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    chinese_units = ['', '拾', '佰', '仟']
    chinese_big_units = ['', '万', '亿', '兆']

    def convert_integer(n):
        if n == 0:
            return '零'
        digits = list(map(int, str(n)))
        length = len(digits)
        result = ''
        zero_flag = False
        for i, d in enumerate(digits):
            unit_index = length - i - 1
            if d == 0:
                zero_flag = True
            else:
                if zero_flag:
                    result += '零'
                    zero_flag = False
                result += chinese_digits[d] + chinese_units[unit_index % 4]
            if unit_index % 4 == 0 and unit_index != 0:
                result += chinese_big_units[unit_index // 4]
        return result

    def convert_decimal(n):
        if n == 0:
            return ''
        if n < 10:
            return chinese_digits[n] + '角'
        else:
            jiao = n // 10
            fen = n % 10
            result = ''
            if jiao > 0:
                result += chinese_digits[jiao] + '角'
            if fen > 0:
                result += chinese_digits[fen] + '分'
            return result

    integer_str = convert_integer(integer_part) if integer_part != 0 else '零'
    decimal_str = convert_decimal(decimal_part)

    # 拼接金额大写
    if not decimal_str:
        result = f'{integer_str}元整'
    else:
        result = f'{integer_str}元{decimal_str}'

    # 如果是负数且金额不为零，在最前面加“负”
    if is_negative and num != 0:
        result = '负' + result

    return result

# ---------- 交互输入签名和起始序号 ----------
print("请设置存货单签名（直接回车表示留空）：")
signature_account = input("请输入记账人（将填入B11）：").strip()
signature_keeper = input("请输入保管人（将填入D11）：").strip()

try:
    start_no = int(input("请输入起始序号（例如1，将生成为10位编号 NO: 0000000001）：").strip() or "1")
except:
    start_no = 1

# ---------- 读取开票明细 ----------
try:
    df = pd.read_excel('开票明细.xlsx', header=0)
except FileNotFoundError:
    print("错误：未找到 '开票明细.xlsx' 文件，请将其放在当前文件夹。")
    input("按回车键退出...")
    exit()

df = df.dropna(how='all').reset_index(drop=True)

# ---------- 按数电发票号码分组，合并数据 ----------
df['数电发票号码'] = df['数电发票号码'].fillna('').astype(str)
grouped = df.groupby('数电发票号码')

merged_rows = []
for invoice, group in grouped:
    if invoice == '':
        continue
    first = group.iloc[0]
    total_quantity = group['数量'].sum()
    total_amount = group['价税合计'].sum()
    avg_price = total_amount / total_quantity if total_quantity != 0 else 0

    merged_row = {
        '门市部': first['门市部'],
        '数电发票号码': invoice,
        '购买方名称': first['购买方名称'],
        '开票日期': first['开票日期'],
        '货物或应税劳务名称': first['货物或应税劳务名称'],
        '单位': first['单位'],
        '数量': total_quantity,
        '单价': avg_price,
        '价税合计': total_amount
    }
    merged_rows.append(merged_row)

merged_df = pd.DataFrame(merged_rows)
total_rows = len(merged_df)

if total_rows == 0:
    print("错误：没有有效的发票数据。")
    input("按回车键退出...")
    exit()

print(f"合并后共有 {total_rows} 张存货单（按发票号分组）。")

# 模板文件路径
template_path = '出库单.xlsx'
if not os.path.exists(template_path):
    print("错误：未找到 '出库单.xlsx' 模板文件。")
    input("按回车键退出...")
    exit()

wb_template = openpyxl.load_workbook(template_path)
ws_template = wb_template.active

wb_result = openpyxl.Workbook()
ws_result = wb_result.active
ws_result.title = '存货单'

# 复制模板列宽
for col in ws_template.column_dimensions:
    ws_result.column_dimensions[col].width = ws_template.column_dimensions[col].width

SLIP_ROWS = 11
EMPTY_ROWS = 5
TOTAL_SLIP_SPAN = SLIP_ROWS + EMPTY_ROWS

left_center = Alignment(horizontal='left', vertical='center')
right_center = Alignment(horizontal='right', vertical='center')
center_center = Alignment(horizontal='center', vertical='center')

# ---------- 遍历生成每张存货单（带错误捕获） ----------
error_invoices = []  # 收集出错的发票号

for idx, row in merged_df.iterrows():
    current_no = start_no + idx
    invoice = str(row.get('数电发票号码', ''))
    print(f'正在生成第 {idx+1}/{total_rows} 张存货单 (序号 NO: {current_no:010d}, 发票: {invoice})...')
    start_row = 1 + idx * TOTAL_SLIP_SPAN

    # 提取数据
    receiver_b3 = str(row.get('购买方名称', '')) if pd.notna(row.get('购买方名称', '')) else ''
    receiver_i11 = str(row.get('门市部', '')) if pd.notna(row.get('门市部', '')) else ''
    date_val = row.get('开票日期', '')
    date_str = pd.to_datetime(date_val).strftime('%Y年%m月%d日') if pd.notna(date_val) else ''
    goods_name_raw = str(row.get('货物或应税劳务名称', '')) if pd.notna(row.get('货物或应税劳务名称', '')) else ''
    goods_name_clean = goods_name_raw.replace('*纺织产品*', '')
    unit = str(row.get('单位', '')) if pd.notna(row.get('单位', '')) else ''
    quantity = row.get('数量', 0)
    unit_price = row.get('单价', 0)
    total_amount = row.get('价税合计', 0)

    # ---- 尝试转换中文大写，失败则记录错误并填充提示文字 ----
    try:
        amount_chinese = digit_to_chinese(total_amount)
    except Exception as e:
        print(f'\n  ⚠️ 错误：发票 {invoice} 的价税合计无法转换为大写！')
        print(f'     金额原始值: {repr(total_amount)}，类型: {type(total_amount).__name__}')
        print(f'     异常信息: {e}')
        error_invoices.append(invoice)
        amount_chinese = '【金额数据错误，请检查原始发票】'

    # 复制模板行和样式
    for r in range(1, SLIP_ROWS + 1):
        target_row = start_row + r - 1
        if ws_template.row_dimensions[r].height is not None:
            ws_result.row_dimensions[target_row].height = ws_template.row_dimensions[r].height
        ws_result.row_dimensions[target_row].hidden = ws_template.row_dimensions[r].hidden

        for c in range(1, 10):
            src = ws_template.cell(row=r, column=c)
            tgt = ws_result.cell(row=target_row, column=c)
            tgt.value = src.value
            if src.has_style:
                tgt.font = copy(src.font)
                tgt.border = copy(src.border)
                tgt.fill = copy(src.fill)
                tgt.number_format = src.number_format
                tgt.protection = copy(src.protection)
                tgt.alignment = copy(src.alignment)

    # 复制模板中的合并单元格
    for merged_range in ws_template.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if max_row <= SLIP_ROWS:
            new_min_row = min_row + start_row - 1
            new_max_row = max_row + start_row - 1
            new_range = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
            ws_result.merge_cells(new_range)

    # ----- 处理第二行：出库单标题和序号 -----
    second_row = start_row + 1
    # 合并 A2:G2 标题
    for cr in ws_result.merged_cells.ranges:
        if (cr.min_row <= second_row <= cr.max_row and cr.min_col <= 7):
            ws_result.unmerge_cells(str(cr))
            break
    ws_result.merge_cells(start_row=second_row, start_column=1, end_row=second_row, end_column=7)
    title_cell = ws_result.cell(row=second_row, column=1)
    title_cell.value = "出库单"
    title_cell.alignment = center_center

    # 处理 H2:I2 序号（10位）
    for cr in ws_result.merged_cells.ranges:
        if (cr.min_row <= second_row <= cr.max_row and 
            cr.min_col <= 8 <= cr.max_col and 
            cr.min_col <= 9 <= cr.max_col):
            ws_result.unmerge_cells(str(cr))
            break
    ws_result.cell(row=second_row, column=8).value = f"NO: {current_no:010d}"
    ws_result.merge_cells(start_row=second_row, start_column=8, end_row=second_row, end_column=9)
    ws_result.cell(row=second_row, column=8).alignment = left_center

    # ----- 填充其他数据 -----
    # 领货单位（B3）
    ws_result.cell(row=start_row+2, column=2).value = receiver_b3
    ws_result.cell(row=start_row+2, column=2).alignment = left_center

    # 日期（H3）
    ws_result.cell(row=start_row+2, column=8).value = date_str
    ws_result.cell(row=start_row+2, column=8).alignment = left_center

    # 明细行（第5行）
    row5 = start_row + 4
    ws_result.cell(row=row5, column=1).value = ''
    ws_result.cell(row=row5, column=2).value = ''
    cell_c5 = ws_result.cell(row=row5, column=3)
    cell_c5.value = goods_name_clean
    cell_c5.alignment = left_center
    ws_result.cell(row=row5, column=4).value = ''
    cell_e5 = ws_result.cell(row=row5, column=5)
    cell_e5.value = unit
    cell_e5.alignment = left_center
    cell_f5 = ws_result.cell(row=row5, column=6)
    cell_f5.value = quantity
    cell_f5.alignment = right_center
    cell_g5 = ws_result.cell(row=row5, column=7)
    cell_g5.value = unit_price
    cell_g5.alignment = right_center
    cell_h5 = ws_result.cell(row=row5, column=8)
    cell_h5.value = total_amount
    cell_h5.alignment = right_center
    ws_result.cell(row=row5, column=9).value = ''

    # 合计行（第9行）
    row9 = start_row + 8
    ws_result.cell(row=row9, column=6).value = quantity
    ws_result.cell(row=row9, column=8).value = total_amount
    for col in range(1, 10):
        ws_result.cell(row=row9, column=col).alignment = right_center

    # 金额行（第10行）
    row10 = start_row + 9
    # 使用捕获后的中文大写（或错误提示）
    ws_result.cell(row=row10, column=2).value = amount_chinese
    ws_result.cell(row=row10, column=6).value = total_amount
    for col in range(1, 10):
        ws_result.cell(row=row10, column=col).alignment = left_center

    # 签名行（第11行）
    row11 = start_row + 10
    cell_b11 = ws_result.cell(row=row11, column=2)
    cell_b11.value = signature_account if signature_account else ''
    cell_b11.alignment = left_center
    cell_d11 = ws_result.cell(row=row11, column=4)
    cell_d11.value = signature_keeper if signature_keeper else ''
    cell_d11.alignment = left_center
    cell_i11 = ws_result.cell(row=row11, column=9)
    cell_i11.value = receiver_i11
    cell_i11.alignment = left_center

# 保存
output_filename = '合并出库单.xlsx'
wb_result.save(output_filename)
print(f'\n全部完成！共生成 {total_rows} 张存货单，已合并保存为：{output_filename}')

# 汇总错误发票信息
if error_invoices:
    print('\n====== 以下发票金额数据异常，大写栏已填入提示文字 ======')
    for inv in error_invoices:
        print(f'-- {inv}')
    print('请检查开票明细.xlsx中对应发票的“价税合计”列。')
else:
    print('所有发票金额转换正常。')

input('按回车键退出...')