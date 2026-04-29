import os
import pandas as pd
import csv
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# ========= 文件名 =========
csv_file = "销售订单列表.csv"
excel_file = "销售订单列表.xlsx"
output_file = "销售订单列表_合并汇总.xlsx"

# ========= 优先选择 CSV =========
if os.path.exists(csv_file):
    input_type = "csv"
    input_file = csv_file
elif os.path.exists(excel_file):
    input_type = "excel"
    input_file = excel_file
else:
    print("❌ 未找到 销售订单列表.csv 或 销售订单列表.xlsx")
    input("按回车退出...")
    exit()

print(f"✅ 已找到文件：{input_file}")
print("开始读取数据...\n")

# ========= 固定列名 =========
NAME_COL = "存货名称"
SPEC_COL = "规格型号"
QTY_COL = "数量"
AMT_COL = "价税合计"

agg = defaultdict(lambda: [0, 0])
total_rows = 0


def to_num(val):
    if val is None:
        return 0
    try:
        s = str(val).replace(",", "").replace("¥", "").replace("￥", "").strip()
        if s == "":
            return 0
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        # 保留原始数值类型（自动判断整数或小数）
        num = float(s)
        if num.is_integer():
            return int(num)
        return num
    except:
        return 0


# ========= 处理 CSV =========
if input_type == "csv":
    # 尝试多种编码
    encodings_to_try = ['utf-8-sig', 'gbk', 'gb2312', 'latin-1']
    reader = None
    f = None
    used_encoding = None
    for enc in encodings_to_try:
        try:
            f = open(input_file, newline='', encoding=enc)
            reader = csv.DictReader(f)
            # 尝试读取表头，触发解码
            _ = reader.fieldnames
            used_encoding = enc
            print(f"✅ 使用编码 {enc} 成功读取")
            break
        except UnicodeDecodeError:
            if f:
                f.close()
            continue
    if reader is None:
        print("❌ 无法用常见编码读取 CSV 文件，请检查文件编码")
        input("按回车退出...")
        exit()

    # 检查必要列
    for col in [NAME_COL, SPEC_COL, QTY_COL, AMT_COL]:
        if col not in reader.fieldnames:
            print(f"❌ 缺少必要列：{col}")
            f.close()
            input("按回车退出...")
            exit()

    print("✅ 表头识别成功")
    print("开始合并处理...\n")

    for row in reader:
        total_rows += 1
        name = row[NAME_COL].strip()
        spec = row[SPEC_COL].strip()
        qty = to_num(row[QTY_COL])
        amt = to_num(row[AMT_COL])

        agg[(name, spec)][0] += qty
        agg[(name, spec)][1] += amt

        if total_rows % 50000 == 0:
            print(f"已处理 {total_rows} 行...")
    f.close()

# ========= 处理 Excel =========
else:
    df = pd.read_excel(input_file)

    for col in [NAME_COL, SPEC_COL, QTY_COL, AMT_COL]:
        if col not in df.columns:
            print(f"❌ 缺少必要列：{col}")
            input("按回车退出...")
            exit()

    print("✅ 表头识别成功")
    print("开始合并处理...\n")

    for _, row in df.iterrows():
        total_rows += 1

        name = str(row[NAME_COL]).strip()
        spec = str(row[SPEC_COL]).strip()
        qty = to_num(row[QTY_COL])
        amt = to_num(row[AMT_COL])

        agg[(name, spec)][0] += qty
        agg[(name, spec)][1] += amt

        if total_rows % 50000 == 0:
            print(f"已处理 {total_rows} 行...")


# ========= 输出结果 =========
print("\n数据处理完成")
print(f"总读取行数：{total_rows}")
print(f"合并后数据条数：{len(agg)}")

# 生成输出列表时，过滤掉“存货名称”和“规格型号”同时为空的行
rows_out = [
    {
        "存货名称": k[0],
        "规格型号": k[1],
        "数量": v[0],
        "价税合计": v[1],
    }
    for k, v in agg.items()
    if k[0] != "" or k[1] != ""   # 至少有一个非空才保留
]

result_df = pd.DataFrame(rows_out)
result_df.sort_values(["存货名称", "规格型号"], inplace=True)

result_df.to_excel(output_file, index=False)

# ========= Excel 格式优化 =========
wb = load_workbook(output_file)
ws = wb.active

# 表头加粗 + 居中
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 冻结首行
ws.freeze_panes = "A2"

# 开启筛选
ws.auto_filter.ref = ws.dimensions

# 自动列宽
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[column].width = max_length + 4

# 数字右对齐
for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
    for cell in row:
        cell.alignment = Alignment(horizontal="right")

wb.save(output_file)

print(f"\n已生成文件：{output_file}")
print(f"实际输出行数：{len(result_df)}")
input("按回车退出...")